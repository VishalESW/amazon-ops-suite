"""Amazon Sponsored Products bulk-upload sheet generator.

Turns the assembled campaign plan into Amazon's Entity-row bulk format (the
"Sponsored Products Campaigns" sheet) so the output uploads directly in Bulk
Operations — no manual conversion. Net-new only: every row Operation=Create with
temp negative IDs linking children to their parent in the same upload, so each
campaign gets exactly one ad group (no duplicates).

Scope: Sponsored Products only — SKW/MKW (keyword), PT/STPP (product targeting),
SPA (auto). Sponsored Brands / Display need creative assets Amazon won't accept
via bulk-create and are intentionally excluded. Negatives are omitted for now.
"""

from __future__ import annotations

import datetime as _dt
import re

import openpyxl

from utils import campaign_orchestrator as orch

# Exact Amazon Sponsored Products bulksheet header order.
SP_HEADERS = [
    "Product", "Entity", "Operation", "Campaign ID", "Ad Group ID", "Portfolio ID",
    "Ad ID", "Keyword ID", "Product Targeting ID", "Campaign Name", "Ad Group Name",
    "Start Date", "End Date", "Targeting Type", "State", "Daily Budget", "SKU",
    "Ad Group Default Bid", "Bid", "Keyword Text", "Native Language Keyword",
    "Native Language Locale", "Match Type", "Bidding Strategy", "Placement",
    "Percentage", "Product Targeting Expression", "Audience ID",
    "Shopper Cohort Percentage", "Shopper Cohort Type", "Sites", "Off-Amazon ad serving",
]

# Sponsored Brands sheet header order (Amazon template).
SB_HEADERS = [
    "Product", "Entity", "Operation", "Campaign ID", "Draft Campaign ID",
    "Portfolio ID", "Ad Group ID", "Keyword ID", "Product Targeting ID",
    "Campaign Name", "Start Date", "End Date", "State", "Budget Type", "Budget",
    "Bid Optimization", "Bid Multiplier", "Bid", "Keyword Text", "Match Type",
    "Product Targeting Expression", "Ad Format", "Landing Page URL",
    "Landing Page ASINs", "Brand Entity ID", "Brand Name", "Brand Logo Asset ID",
    "Custom Image Asset ID", "Creative Headline", "Creative ASINs", "Video Media IDs",
    "Creative Type",
]

# Sponsored Display sheet header order (Amazon template).
SD_HEADERS = [
    "Product", "Entity", "Operation", "Campaign ID", "Portfolio ID", "Ad Group ID",
    "Ad ID", "Targeting ID", "Campaign Name", "Ad Group Name", "Start Date",
    "End Date", "State", "Tactic", "Budget Type", "Budget", "SKU",
    "Ad Group Default Bid", "Bid", "Bid Optimization", "Cost Type",
    "Targeting Expression",
]

# Planning Match Type -> Amazon keyword match type.
_MATCH = {"Ex.": "exact", "Br.": "broad", "Br.M": "broad", "Ph.": "phrase"}
# Planning Bidding Strategy -> Amazon strategy.
_STRATEGY = {"Fixed Bids": "Fixed bid", "Down Only": "Dynamic bids - down only",
             "Up and Down Only": "Dynamic bids - up and down"}
# PAT category -> BuildInput ASIN-bucket attribute.
_PAT_BUCKET = {"Main": "main_competitor_asins", "Low Rated": "lower_rated_asins",
               "High Priced": "higher_priced_asins", "Bestselling": "bestselling_asins"}

DEFAULT_AG_BID = 0.75


def _clean(terms):
    """Drop junk (blanks, 1-char, flag values like 'Y'/'N', non-alpha) + dedupe."""
    out = []
    for t in terms:
        t = str(t or "").strip()
        if len(t) < 2 or not any(ch.isalpha() for ch in t):
            continue
        out.append(t)
    return list(dict.fromkeys(out))


def _le4(terms):
    """Negative PHRASE keywords are capped at 4 words by Amazon — drop longer ones."""
    return [t for t in terms if t and len(str(t).split()) <= 4]


def _negatives(cr, inp, skw_kws):
    """Negative keyword/product rows for a campaign. Negative PHRASE + EXACT terms are
    sourced from the Master Keyword List sheet:
      - negativePhrase  = own-branded (AL/AM) + competitor searches (AQ) + Negate Brands (BB)
      - negativeExact   = Negate Words (BA)
      - negative product = own-brand ASINs (BD)
    Gated by the planning negative LABELS in Campaign Naming (AC/AD/AE) so only the
    campaigns marked for negatives receive them."""
    own_kw = _le4(_clean((getattr(inp, "own_branded_kws", None) or [])
                         + (getattr(inp, "own_branded_searches", None) or [])))
    comp_kw = _le4(_clean(getattr(inp, "competitor_searches", None) or []))
    neg_brands = _le4(_clean(getattr(inp, "negate_brands", None) or []))
    neg_words = _clean(getattr(inp, "negate_words", None) or [])
    out = []
    ac = str(cr.get("AC") or "")
    # A literal "None" label means NO negatives — don't treat it as a real label.
    has_phrase = ("Own Branded KWs" in ac) or ("Competitor KWs" in ac)
    has_exact = "SKW EX. Match Keywords" in str(cr.get("AD") or "")
    if "Own Branded KWs" in ac:
        out += [("Negative Keyword", k, "negativePhrase") for k in own_kw]
    if "Competitor KWs" in ac:
        out += [("Negative Keyword", k, "negativePhrase") for k in comp_kw]
    if has_phrase:   # a real phrase-negative label -> also add Master List Negate Brands
        out += [("Negative Keyword", k, "negativePhrase") for k in neg_brands]
    if has_exact:    # exact-negative label -> Master List Negate Words
        out += [("Negative Keyword", k, "negativeExact") for k in neg_words]
    if "Own Branded ASINs" in str(cr.get("AE") or ""):
        out += [("Negative Product Targeting", a, "")
                for a in (getattr(inp, "own_brand_asins", None) or [])]
    # Self-negation: the SKW-exact keywords are added as NEGATIVE EXACT in the broader
    # MKW campaigns (Phrase / Broad / Broad-modifier) so those match types don't
    # cannibalize the dedicated SKW-exact campaigns.
    if str(cr.get("E") or "").upper() == "MKW" and str(cr.get("F") or "") in ("Br.", "Br.M", "Ph."):
        out += [("Negative Keyword", k, "negativeExact") for k in _clean(skw_kws)]
    # de-dupe (entity, value, match)
    return list(dict.fromkeys(out))


def _split_broad(text):
    """Broad KW List cell -> list of keyword lines (newline/comma separated)."""
    if not text:
        return []
    parts = re.split(r"[\n,]+", str(text))
    return [p.strip() for p in parts if p and p.strip()]


def _num(v):
    return v if isinstance(v, (int, float)) and not isinstance(v, bool) else None


def _placement_pct(v):
    """Campaign-Naming placement adjustment (fraction like 0.25/0.5/1.0 or a whole
    percent) -> Amazon Percentage integer. Returns None for 0/blank."""
    n = _num(v)
    if n is None:
        try:
            n = float(str(v).replace("%", "").strip())
        except (TypeError, ValueError):
            return None
    if not n:
        return None
    return round(n * 100) if abs(n) <= 10 else round(n)


# Campaign-Naming placement column -> Amazon placement name.
_PLACEMENTS = [("Q", "Placement Top"), ("R", "Placement Product Page"),
               ("S", "Placement Rest Of Search")]


def _name(cr):
    """Campaign Name = TEXTJOIN(' | ', skip blanks, B,C,D,E,F,G,H)."""
    parts = [cr.get(c, "") for c in ("B", "C", "D", "E", "F", "G", "H")]
    return " | ".join(str(p) for p in parts if p not in (None, ""))


def _ag_name(cr):
    """Ad Group Name = TEXTJOIN(' | ', skip blanks, B,E,G,F)."""
    parts = [cr.get(c, "") for c in ("B", "E", "G", "F")]
    return " | ".join(str(p) for p in parts if p not in (None, ""))


def build_sp_rows(inp):
    """Return a list of Amazon SP bulksheet row-dicts (keyed by SP_HEADERS) for
    every Sponsored Products campaign in the plan."""
    own_asin = inp.products[0]["asin"] if inp.products else ""
    own_sku = inp.products[0].get("sku", "") if inp.products else ""
    today = _dt.date.today().strftime("%Y%m%d")

    # Keyword pool per (root, match) for MKW expansion.
    sem = inp.semantics_rows
    # All SKW keywords (exact) — used to expand the "SKW EX. Match Keywords" negative.
    skw_kws = [s["keyword"] for s in sem if (s.get("disp_kw_type") or "").upper() == "SKW"]
    rows = []
    cid = -1   # temp campaign id counter (negative)

    def add(**kw):
        r = {h: "" for h in SP_HEADERS}
        r["Product"] = "Sponsored Products"
        r["Operation"] = "Create"
        r.update(kw)
        rows.append(r)

    for cr in inp.campaign_rows:
        ctype, e, f, g = cr.get("C"), cr.get("E"), cr.get("F"), cr.get("G")
        if ctype not in ("SPM", "SPA"):
            continue  # SB/SBV/SDI handled elsewhere (creative assets required)
        camp_id, ag_id = cid, cid       # reuse the same temp id within a campaign
        cid -= 1
        is_auto = ctype == "SPA"
        strategy = _STRATEGY.get(cr.get("P"), "Dynamic bids - down only")
        # Campaign
        add(Entity="Campaign", **{"Campaign ID": camp_id, "Campaign Name": _name(cr),
            "Start Date": today, "Targeting Type": "AUTO" if is_auto else "MANUAL",
            "State": "enabled", "Daily Budget": cr.get("K", 5) or 5,
            "Bidding Strategy": strategy})
        # Campaign placement bid adjustments (Top / Product Page / Rest of Search)
        # from Campaign Naming Q/R/S -> Amazon "Bidding Adjustment" rows.
        for col, pname in _PLACEMENTS:
            p = _placement_pct(cr.get(col))
            if p:
                add(Entity="Bidding Adjustment", **{"Campaign ID": camp_id,
                    "Bidding Strategy": strategy, "Placement": pname, "Percentage": p})
        # Ad Group. Default bid = Starting Bid (AL, computed) if present, else Z, else default.
        ag_bid = _num(cr.get("AL")) or _num(cr.get("Z")) or DEFAULT_AG_BID
        add(Entity="Ad Group", **{"Campaign ID": camp_id, "Ad Group ID": ag_id,
            "Ad Group Name": _ag_name(cr), "State": "enabled",
            "Ad Group Default Bid": ag_bid})
        # Product Ad (the advertised own product)
        if own_sku or own_asin:
            add(Entity="Product Ad", **{"Campaign ID": camp_id, "Ad Group ID": ag_id,
                "State": "enabled", "SKU": own_sku or own_asin})
        # Negatives (expanded from the planning AC/AD/AE labels)
        for entity, val, match in _negatives(cr, inp, skw_kws):
            if entity == "Negative Keyword":
                add(Entity="Negative Keyword", **{"Campaign ID": camp_id, "Ad Group ID": ag_id,
                    "State": "enabled", "Keyword Text": val, "Match Type": match})
            else:  # Negative Product Targeting
                add(Entity="Negative Product Targeting", **{"Campaign ID": camp_id,
                    "Ad Group ID": ag_id, "State": "enabled",
                    "Product Targeting Expression": f'asin="{val}"'})
        # Targeting
        if is_auto:
            continue  # auto campaigns: Amazon creates the auto-targeting groups
        if e in ("SKW", "MKW"):
            match = _MATCH.get(f, "exact")
            # Targeting keywords come from the Semantics "Broad KW List" (col O);
            # fall back to the raw keyword(s) when the user left it blank.
            # Each keyword's bid = its own Semantics "Starting Bid" (per-keyword CVR),
            # falling back to the ad-group default only when that's blank. Avoids the
            # flat ~1.17 that Campaign Naming's placeholder-CVR Starting Bid produced.
            if e == "SKW":
                srow = next((s for s in sem if s.get("keyword") == g
                             and (s.get("disp_kw_type") or "").upper() == "SKW"), None)
                kbid = (srow.get("bid") if srow else None) or ag_bid
                texts = (_split_broad(srow.get("broad_list")) if srow else []) or [g]
                kws = [(t, kbid) for t in dict.fromkeys(texts)]
            else:  # MKW: rows under this root + match — bid per source keyword
                group = [s for s in sem
                         if (s.get("category") or "").strip() == g
                         and (s.get("disp_kw_type") or "").upper() == "MKW"
                         and orch.canon_match(s.get("disp_match")) == f]
                kws, seen = [], set()
                for s in group:
                    kbid = s.get("bid") or ag_bid
                    for t in (_split_broad(s.get("broad_list")) or [s["keyword"]]):
                        if t and t not in seen:
                            seen.add(t)
                            kws.append((t, kbid))
            for text, bid in kws:
                add(Entity="Keyword", **{"Campaign ID": camp_id, "Ad Group ID": ag_id,
                    "State": "enabled", "Keyword Text": text, "Match Type": match,
                    # explicit bid on every keyword: its own, else the ad-group bid.
                    "Bid": bid if isinstance(bid, (int, float)) and bid else ag_bid})
        elif e == "STPP":
            # Brand defence: target the OWN-BRANDED keywords (from Master KW List),
            # exact match — so competitors can't win the brand's own terms.
            match = _MATCH.get(f, "exact")
            own_kws = _clean((getattr(inp, "own_branded_kws", None) or [])
                             + (getattr(inp, "own_branded_searches", None) or []))
            for text in own_kws:
                add(Entity="Keyword", **{"Campaign ID": camp_id, "Ad Group ID": ag_id,
                    "State": "enabled", "Keyword Text": text, "Match Type": match,
                    "Bid": ag_bid})
        elif e == "PT":
            # PT targets the competitor ASINs of the category (G) via product targeting.
            for asin in (getattr(inp, _PAT_BUCKET.get(g, ""), []) or []):
                add(Entity="Product Targeting", **{"Campaign ID": camp_id,
                    "Ad Group ID": ag_id, "State": "enabled", "Bid": ag_bid,
                    "Product Targeting Expression": f'asin="{asin}"'})
    return rows


def _brand_name(inp):
    """Brand Name for Sponsored Brands: the Ads profile (Campaign Naming col J),
    else the advertised product's name."""
    for cr in (getattr(inp, "campaign_rows", None) or []):
        j = str(cr.get("J") or "").strip()
        if j:
            return j
    return inp.products[0].get("name", "") if getattr(inp, "products", None) else ""


def _root_keywords(sem, root):
    """Broad KW List keywords for a Root KW group (fallback to the raw keywords)."""
    texts = []
    for s in sem:
        if (s.get("category") or "").strip() == root and \
                (s.get("disp_kw_type") or "").upper() in ("SKW", "MKW"):
            texts += _split_broad(s.get("broad_list")) or ([s["keyword"]] if s.get("keyword") else [])
    return [t for t in dict.fromkeys(texts) if t]


def build_sb_rows(inp):
    """Sponsored Brands rows (SB = Product Collection, SBV = Video) from the SB/SBV
    campaign-plan rows. Keywords come from the root's Broad KW List. Account-only
    creative fields (Brand Entity ID, Brand Logo Asset ID, Creative Headline, Video
    Media IDs) are left blank for the user to fill from their Amazon asset library."""
    own_asin = inp.products[0]["asin"] if inp.products else ""
    brand = _brand_name(inp)
    today = _dt.date.today().strftime("%Y%m%d")
    sem = inp.semantics_rows
    rows, cid = [], -1

    def add(**kw):
        r = {h: "" for h in SB_HEADERS}
        r["Product"], r["Operation"] = "Sponsored Brands", "Create"
        r.update(kw)
        rows.append(r)

    for cr in inp.campaign_rows:
        if cr.get("C") not in ("SB", "SBV"):
            continue
        is_video = cr.get("C") == "SBV"
        camp_id = cid
        cid -= 1
        texts = _root_keywords(sem, cr.get("G")) or ([cr.get("G")] if cr.get("G") else [])
        bid = _num(cr.get("AL")) or _num(cr.get("Z")) or DEFAULT_AG_BID
        add(Entity="Campaign", **{"Campaign ID": camp_id, "Campaign Name": _name(cr),
            "Start Date": today, "State": "enabled", "Budget Type": "Daily",
            "Budget": cr.get("K", 5) or 5, "Bid Optimization": "On",
            "Ad Format": "Video" if is_video else "Product Collection",
            "Landing Page ASINs": own_asin, "Brand Name": brand,
            "Creative ASINs": own_asin,
            "Creative Type": "Video" if is_video else "Product Collection"})
        for t in texts:
            add(Entity="Keyword", **{"Campaign ID": camp_id, "Keyword Text": t,
                "Match Type": _MATCH.get(cr.get("F"), "exact"), "Bid": bid})
    return rows


def build_sd_rows(inp):
    """Sponsored Display rows (SDI). Tactic = Views (CPC / T00020), product targeting.
    SDI PT targets the competitor ASINs of its category; SDI remarketing (VREM/PREM)
    targets the own product's ASIN (best-effort — audience remarketing expressions are
    account-specific and left for the user to refine)."""
    own_asin = inp.products[0]["asin"] if inp.products else ""
    own_sku = inp.products[0].get("sku", "") if inp.products else ""
    today = _dt.date.today().strftime("%Y%m%d")
    rows, cid = [], -1

    def add(**kw):
        r = {h: "" for h in SD_HEADERS}
        r["Product"], r["Operation"] = "Sponsored Display", "Create"
        r.update(kw)
        rows.append(r)

    for cr in inp.campaign_rows:
        if cr.get("C") != "SDI":
            continue
        camp_id = ag_id = cid
        cid -= 1
        bid = _num(cr.get("AL")) or _num(cr.get("Z")) or DEFAULT_AG_BID
        add(Entity="Campaign", **{"Campaign ID": camp_id, "Campaign Name": _name(cr),
            "Start Date": today, "State": "enabled", "Tactic": "T00020",
            "Budget Type": "Daily", "Budget": cr.get("K", 5) or 5,
            "Bid Optimization": "clicks", "Cost Type": "cpc"})
        add(Entity="Ad Group", **{"Campaign ID": camp_id, "Ad Group ID": ag_id,
            "Ad Group Name": _ag_name(cr), "State": "enabled",
            "Ad Group Default Bid": bid})
        if own_sku or own_asin:
            add(Entity="Product Ad", **{"Campaign ID": camp_id, "Ad Group ID": ag_id,
                "State": "enabled", "SKU": own_sku or own_asin})
        # Targeting
        if cr.get("E") == "PT":
            asins = list(getattr(inp, _PAT_BUCKET.get(cr.get("G"), ""), []) or [])
        else:  # VREM / PREM remarketing -> own product (best-effort)
            asins = [own_asin] if own_asin else []
        for asin in asins:
            add(Entity="Product Targeting", **{"Campaign ID": camp_id, "Ad Group ID": ag_id,
                "State": "enabled", "Bid": bid, "Cost Type": "cpc",
                "Targeting Expression": f'asin="{asin}"'})
    return rows


def build_sp_bulksheet(inp, out_path):
    """Write the Amazon bulksheet .xlsx (SP + SB + SD sheets). Returns (path, n_campaigns)."""
    sp, sb, sd = build_sp_rows(inp), build_sb_rows(inp), build_sd_rows(inp)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sponsored Products Campaigns"
    ws.append(SP_HEADERS)
    for r in sp:
        ws.append([r.get(h, "") for h in SP_HEADERS])
    wsb = wb.create_sheet("Sponsored Brands Campaigns")
    wsb.append(SB_HEADERS)
    for r in sb:
        wsb.append([r.get(h, "") for h in SB_HEADERS])
    wsd = wb.create_sheet("Sponsored Display Campaigns")
    wsd.append(SD_HEADERS)
    for r in sd:
        wsd.append([r.get(h, "") for h in SD_HEADERS])
    wb.save(out_path)
    n_camp = sum(1 for r in (sp + sb + sd) if r["Entity"] == "Campaign")
    return out_path, n_camp


# --------------------------------------------------------------------------- #
# Build from a user-edited planning workbook (Verify & Build "upload edited")  #
# --------------------------------------------------------------------------- #
_ASIN_ONLY = re.compile(r"^B0[A-Z0-9]{8}$", re.I)
_PAT_TAG_BUCKET = {"main": "main_competitor_asins", "low rated": "lower_rated_asins",
                   "high priced": "higher_priced_asins", "bestselling": "bestselling_asins"}


def _col_idx(header, *names):
    for i, h in enumerate(header):
        hl = str(h or "").strip().lower()
        if hl in [n.lower() for n in names]:
            return i
    return None


def parse_planning_workbook(path):
    """Re-read the app's planning workbook (possibly hand-edited) into the object
    build_sp_rows expects: campaign_rows (letter-keyed dicts from Campaign Naming),
    semantics_rows (for MKW keyword expansion), the four PAT ASIN buckets, and the
    own ASIN/SKU. Lets the user tweak the workbook, then regenerate the bulksheet."""
    from openpyxl.utils import get_column_letter
    from types import SimpleNamespace
    wb = openpyxl.load_workbook(path, data_only=True)

    # --- Campaign Naming -> campaign_rows (keyed by column letter, A,B,C…) ------
    campaign_rows, own_asin, own_sku = [], "", ""
    cn = wb["Campaign Naming, Bids & Targets"]
    for r in cn.iter_rows(min_row=2, values_only=True):
        if not any(c not in (None, "") for c in r):
            continue
        cr = {get_column_letter(i + 1): ("" if v is None else v) for i, v in enumerate(r)}
        if str(cr.get("C", "")).strip():          # a real campaign row (has a type)
            campaign_rows.append(cr)
            if not own_asin:
                aa = str(cr.get("AA", "") or "").strip()   # ASIN/SKU column
                if "/" in aa:
                    own_asin, own_sku = (aa.split("/", 1) + [""])[:2]
                elif _ASIN_ONLY.match(aa):
                    own_asin = aa

    # --- Semantics -> keyword rows (Keyword / Root KW / KW Vol. / Match Type) ----
    sem_rows = []
    if "Semantics" in wb.sheetnames:
        sm = list(wb["Semantics"].iter_rows(values_only=True))
        hdr = next((r for r in sm[:6] if r and any("keyword" == str(c).strip().lower()
                    for c in r)), sm[2] if len(sm) > 2 else [])
        h = [str(c or "") for c in hdr]
        i_kw, i_root = _col_idx(h, "Keyword"), _col_idx(h, "Root KW")
        i_kv, i_mt = _col_idx(h, "KW Vol."), _col_idx(h, "Match Type")
        i_broad = _col_idx(h, "Broad KW List")   # Semantics col O — targeting keywords
        i_bid = _col_idx(h, "Starting Bid 1", "Starting Bid")  # per-keyword bid (col U)
        start = sm.index(hdr) + 1 if hdr in sm else 3
        for r in sm[start:]:
            if not r or i_kw is None or i_kw >= len(r) or not r[i_kw]:
                continue
            g = lambda i: (str(r[i]).strip() if i is not None and i < len(r) and r[i] else "")
            sem_rows.append({"keyword": g(i_kw), "category": g(i_root),
                             "disp_kw_type": g(i_kv), "disp_match": g(i_mt),
                             "broad_list": g(i_broad),
                             "bid": _num(r[i_bid]) if i_bid is not None and i_bid < len(r) else None})

    # --- PAT -> the four competitor-ASIN buckets (by ASIN Cat Type) -------------
    buckets = {b: [] for b in _PAT_TAG_BUCKET.values()}
    if "PAT" in wb.sheetnames:
        pt = list(wb["PAT"].iter_rows(values_only=True))
        phdr = next((r for r in pt[:4] if r and any("cat type" in str(c).strip().lower()
                     for c in r)), pt[1] if len(pt) > 1 else [])
        ph = [str(c or "") for c in phdr]
        i_asin = _col_idx(ph, "Search Term") or 1
        i_cat = _col_idx(ph, "ASIN Cat Type")
        pstart = pt.index(phdr) + 1 if phdr in pt else 2
        for r in pt[pstart:]:
            if not r or i_asin >= len(r):
                continue
            asin = str(r[i_asin] or "").strip()
            cat = str(r[i_cat] or "").strip().lower() if i_cat is not None and i_cat < len(r) else ""
            if _ASIN_ONLY.match(asin) and cat in _PAT_TAG_BUCKET:
                buckets[_PAT_TAG_BUCKET[cat]].append(asin)

    # --- Master Keyword List -> negative-keyword source lists ------------------
    # Engine writes these columns (data from row 3): AL own-branded KWs,
    # AM own-branded searches, AQ competitor searches, BD own-brand ASINs.
    from openpyxl.utils import column_index_from_string
    neg = {"own_branded_kws": "AL", "own_branded_searches": "AM",
           "competitor_searches": "AQ", "own_brand_asins": "BD",
           "negate_words": "BA", "negate_brands": "BB"}   # ⛔️ negatives
    neg_lists = {k: [] for k in neg}
    if "Master Keyword List" in wb.sheetnames:
        mk = list(wb["Master Keyword List"].iter_rows(values_only=True))
        idx = {k: column_index_from_string(v) - 1 for k, v in neg.items()}
        for r in mk[2:]:                                  # data starts row 3
            for k, ci in idx.items():
                if ci < len(r) and r[ci] not in (None, ""):
                    neg_lists[k].append(str(r[ci]).strip())

    return SimpleNamespace(products=[{"asin": own_asin, "sku": own_sku}],
                           semantics_rows=sem_rows, campaign_rows=campaign_rows,
                           **buckets, **neg_lists)


def build_sp_bulksheet_from_workbook(path, out_path):
    """Parse an edited planning workbook and write the Amazon SP bulksheet from it."""
    return build_sp_bulksheet(parse_planning_workbook(path), out_path)
