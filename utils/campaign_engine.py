"""Campaign Processor — workbook builder.

Reproduces the client's 20-tab Helium10 -> Amazon PPC campaign-creation workbook
from user uploads + manual inputs, preserving LIVE formulas.

Strategy: load `assets/campaign_template.xlsx` (an exact copy of the master),
clear example data, write the user's data, and (re)write the formula columns with
exact templates. Cross-sheet XLOOKUPs stay IFERROR-guarded so missing OPTIONAL
tabs (Search Term Report, extra SQP reports) never break the sheet.

See docs/CAMPAIGN_PROCESSOR_SPEC.md for the full reverse-engineered spec.
"""

from __future__ import annotations

import os
import re
from dataclasses import dataclass, field

import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.utils import get_column_letter, column_index_from_string

from config import cfg

TEMPLATE_PATH = os.path.join(cfg.ROOT, "assets", "campaign_template.xlsx")

# Tab names exactly as they appear in the template.
S_ASIN = "ASIN List"
S_POE = "Product Opportunity Explorer"
S_H10 = "Helium10 Reverse ASIN Search"
S_STR = "Search Term Report"
S_BRAND_ANALYTICS = "Brand Analytics - ASIN Filter"
S_BRAND = "Brand"
S_MKL = "Master Keyword List"
S_PAT = "PAT"
S_SEM = "Semantics"
S_CAMP = "Campaign Naming, Bids & Targets"
S_ROOT = "Campaign Root KW"

# A SQP tab is named "SQP Report <ASIN>"; the template ships one example
# ("SQP Report B0DQV3BZD7") that we rename/clone per the user's own ASINs.
TEMPLATE_SQP = "SQP Report B0DQV3BZD7"
# BA TST tabs are named "BA TST - <word>"; the template ships several examples.


# --------------------------------------------------------------------------- #
# Layout: where each sheet's data table lives.
# --------------------------------------------------------------------------- #
@dataclass
class Layout:
    header_row: int          # row holding column labels
    data_row: int            # first row of data
    last_col: str            # rightmost column to manage
    clear_to: int = 1000     # clear example rows down to here


LAYOUTS = {
    S_ASIN: Layout(2, 3, "AB"),
    S_POE:  Layout(22, 23, "Q"),
    S_H10:  Layout(15, 16, "AF", clear_to=3100),
    S_STR:  Layout(6, 7, "BD", clear_to=17500),
    S_BRAND_ANALYTICS: Layout(17, 18, "AB"),
    S_BRAND: Layout(6, 7, "U"),
    S_MKL:  Layout(2, 3, "BO"),
    S_PAT:  Layout(2, 3, "T"),
    S_SEM:  Layout(3, 4, "U"),
    S_CAMP: Layout(1, 2, "AN", clear_to=901),
    TEMPLATE_SQP: Layout(11, 12, "AI"),
    S_ROOT: Layout(1, 2, "A"),
}
# BA TST layout (any tab starting with "BA TST")
BA_LAYOUT = Layout(9, 10, "V", clear_to=7000)
# The H10 / Brand ASIN-input list lives above the keyword table.
H10_ASIN_START = 4   # A4 down, stop before keyword header (row 15)
BRAND_ASIN_START = 4


# --------------------------------------------------------------------------- #
# Formula templates (row r).
# Cross-sheet lookups use INDEX/MATCH (not XLOOKUP): works in every Excel version
# AND Google Sheets, avoids the `_xlfn.` prefix trap that makes bare XLOOKUP read as
# an unknown name (#NAME? -> IFERROR -> 0 / "n/a"). Each is IFERROR-guarded.
# --------------------------------------------------------------------------- #
def _lk(ret, key, look, default):
    """IFERROR(INDEX(ret, MATCH(key, look, 0)), default) — XLOOKUP equivalent."""
    return f"IFERROR(INDEX({ret},MATCH({key},{look},0)),{default})"


def _asin_formulas(r):
    return {
        "D": f'=B{r}&"/"&A{r}',
        "E": f"=SUMIFS(Q:Q,L:L,B{r})",
        "F": f"=SUMIFS(R:R,L:L,B{r})",
        "G": f"=SUMIFS(S:S,L:L,B{r})",
    }


def _pat_formulas(r, has_str):
    str_ = "'Search Term Report'"
    def lk(col):
        return "=" + _lk(f"{str_}!{col}:{col}", f"B{r}", f"{str_}!A:A", "0") if has_str else "0"
    return {
        "D": f'="www.amazon.com/dp/"&B{r}',
        "E": lk("S"),
        "F": lk("U"),
        "G": lk("Q"),
        "H": lk("R"),
        "L": "=" + _lk(f"'{S_CAMP}'!$I:$I", f'K{r}&"-PT-Ex.-"&$J{r}',
                       f"'{S_CAMP}'!$AN:$AN", '"n/a"'),
        "M": "=" + _lk("$T:$T", f"$J{r}", "$S:$S", '""'),
        # Conversion rate as a fraction (CVR sources may be % or fraction).
        "P": f"=IF(H{r}>1,H{r}/100,H{r})",
        "Q": f"=ROUND((N{r}*O{r}*P{r})/(1+M{r}),2)",
    }


def _sem_search_volume_formula(r, sqp_tabs):
    """C: search volume. POE first, then each existing SQP tab, then H10.
    H10 search volume lives in column G ('Search Volume'); H is the trend."""
    inner = _lk(f"'{S_H10}'!G:G", f"A{r}", f"'{S_H10}'!A:A", "0")
    for tab in reversed(sqp_tabs):
        inner = _lk(f"'{tab}'!D:D", f"A{r}", f"'{tab}'!A:A", inner)
    return "=" + _lk(f"'{S_POE}'!P:P", f"A{r}", f"'{S_POE}'!C:C", inner)


# Marker written into Search-Term-Report-dependent columns when that file is absent.
STR_MISSING_NOTE = "⚠ Upload Search Term Report"


def _sem_formulas(r, has_str, sqp_tabs):
    str_ = "'Search Term Report'"
    def lk(col):
        return "=" + _lk(f"{str_}!{col}:{col}", f"A{r}", f"{str_}!A:A", "0") \
            if has_str else STR_MISSING_NOTE
    # Conversion Rate priority: POE "Search Conversion Rate (360 days)" (POE!I by
    # POE!C) -> Search Term Report Orders/Clicks (computed) -> STR "CVR" column
    # (STR!R) -> H10 "ABA Total Conv. Share" (H10!C by H10!A). First source with a
    # value > 0 wins. Amazon STR exports frequently leave the CVR column blank while
    # still carrying Orders (U) and Clicks (P), so we DERIVE CVR = Orders/Clicks
    # from those instead of depending on the blank column. Values may be % or
    # fraction; normalize if > 1.
    poe_cvr = _lk(f"'{S_POE}'!I:I", f"A{r}", f"'{S_POE}'!C:C", "0")
    h10_cvr = _lk(f"'{S_H10}'!C:C", f"A{r}", f"'{S_H10}'!A:A", "0")
    sources = [poe_cvr]
    if has_str:
        str_orders = f"IFERROR(INDEX({str_}!U:U,MATCH(A{r},{str_}!A:A,0)),0)"
        str_clicks = f"IFERROR(INDEX({str_}!P:P,MATCH(A{r},{str_}!A:A,0)),0)"
        sources.append(f"IFERROR(({str_orders})/({str_clicks}),0)")   # Orders/Clicks
        sources.append(_lk(f"{str_}!R:R", f"A{r}", f"{str_}!A:A", "0"))  # direct CVR col
    sources.append(h10_cvr)
    # Nest into IF(src1>0, src1, IF(src2>0, src2, src3)) — first positive wins.
    raw = f"({sources[-1]})"
    for s in reversed(sources[:-1]):
        raw = f"IF(({s})>0,({s}),{raw})"
    cvr = f"=IF(({raw})>1,({raw})/100,({raw}))"
    # Layout: K Root KW, L (blank spacer), M KW Vol., N Match, O Broad, P Product,
    # Q Campaign Name, R Placement Mod, S ASP, T ACoS Target, U Starting Bid.
    return {
        "C": _sem_search_volume_formula(r, sqp_tabs),
        "D": lk("U"),
        "E": lk("S"),
        "F": lk("Q"),
        "G": cvr,
        # Campaign-name lookup key must match Campaign Naming AN (= B-E-F-G):
        # SKW campaigns key on the keyword (A), MKW campaigns on the Root KW (K),
        # since MKW campaigns are grouped per root (G = root).
        "Q": "=" + _lk(f"'{S_CAMP}'!$I:$I",
                       f'P{r}&"-"&$M{r}&"-"&$N{r}&"-"&IF($M{r}="MKW",$K{r},$A{r})',
                       f"'{S_CAMP}'!$AN:$AN", '"n/a"'),
        "U": f"=ROUND((G{r}*S{r}*T{r})/(1+R{r}),2)",
    }


def _camp_formulas(r, brand):
    return {
        # Name joins B|C|D|E|F|G|H, skipping blank cells (TEXTJOIN ignore_empty),
        # so SPA/STPP (no E/G) and SB/SBV/SDI (Landing Page in D) all read correctly.
        "I": f'=_xlfn.TEXTJOIN(" | ",TRUE,B{r},C{r},D{r},E{r},F{r},G{r},H{r})',
        "T": f'="{brand} | "&B{r}&" > "&H{r}',
        "W": f'=_xlfn.TEXTJOIN(" | ",TRUE,B{r},E{r},G{r},F{r})',
        "AN": f'=B{r}&"-"&E{r}&"-"&F{r}&"-"&G{r}',
        "AL": f"=ROUND((AI{r}*AJ{r}*AK{r})/(1+AH{r}),2)",
    }


# --------------------------------------------------------------------------- #
# Low-level helpers
# --------------------------------------------------------------------------- #
def _put(ws, row, col, val):
    """Set a cell value, skipping merged (read-only) cells."""
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        return
    cell.value = val


def _clear(ws, layout: Layout):
    # Template ships clean (no example data), so we only need to clear whatever
    # rows actually exist — iterating beyond max_row would create cells and
    # re-inflate the sheet.
    last = column_index_from_string(layout.last_col)
    end = min(layout.clear_to, ws.max_row)
    for r in range(layout.data_row, end + 1):
        for c in range(1, last + 1):
            _put(ws, r, c, None)


def _write_rows(ws, rows, start_row, start_col=1):
    """Paste a list of row-lists positionally."""
    for i, row in enumerate(rows):
        for j, val in enumerate(row):
            if val is None or val == "":
                continue
            _put(ws, start_row + i, start_col + j, val)


def _set(ws, col, r, val):
    _put(ws, r, column_index_from_string(col), val)


def _setf(ws, col, r, val, fmt):
    """Set value AND number format (so percentages show as 30%, not 0.3)."""
    c = ws.cell(row=r, column=column_index_from_string(col))
    if isinstance(c, MergedCell):
        return
    c.value = val
    c.number_format = fmt


def _set_fmt(ws, r, colmap):
    """Apply number formats to several columns at row r (value untouched)."""
    for col, fmt in colmap.items():
        c = ws.cell(row=r, column=column_index_from_string(col))
        if not isinstance(c, MergedCell):
            c.number_format = fmt


PCT = "0%"
PCT1 = "0.0%"
MONEY = "#,##0.00"
INT = "#,##0"


# --------------------------------------------------------------------------- #
# Public build
# --------------------------------------------------------------------------- #
@dataclass
class BuildInput:
    brand: str = "Brand"                 # profile / brand name used in tags
    products: list = field(default_factory=list)   # [{asin,sku,name}]
    # passthrough tables: {sheet_or_key: [[..row..], ...]} (data rows only, no header)
    poe_tables: dict = field(default_factory=dict)   # {product_name: rows}  cols B..O(+Q)
    h10_asins: list = field(default_factory=list)
    h10_table: list = field(default_factory=list)    # cols A..AF
    str_table: list = field(default_factory=list)    # optional
    brand_analytics_table: list = field(default_factory=list)
    brand_asins: list = field(default_factory=list)
    brand_table: list = field(default_factory=list)
    sqp_reports: dict = field(default_factory=dict)   # {asin: rows}
    ba_tst: dict = field(default_factory=dict)        # {word: rows}  6-block flat per upload
    # manual MKL
    own_branded_kws: list = field(default_factory=list)
    own_branded_searches: list = field(default_factory=list)
    competitor_kws: list = field(default_factory=list)
    competitor_searches: list = field(default_factory=list)
    own_brand_asins: list = field(default_factory=list)
    # competitor ASIN buckets (MKL BE..BI) — also drive PAT targeting
    main_competitor_asins: list = field(default_factory=list)
    lower_rated_asins: list = field(default_factory=list)
    higher_priced_asins: list = field(default_factory=list)
    bestselling_asins: list = field(default_factory=list)
    harvested_asins: list = field(default_factory=list)
    # PAT rows (competitor ASINs we target): {asin,type,title,product,asp,acos,source}
    pat_targets: list = field(default_factory=list)
    # Campaign Root KW categories (col A), client-specific.
    root_categories: list = field(default_factory=list)
    # ASIN List CTR block (cols L..AA), per-ASIN AdLabs metrics. Each dict:
    # {asin,sku,title,state,profile,impression,click,ctr,spend,cpc,sales,orders,acos,price,asp}
    asin_ctr: list = field(default_factory=list)
    # Semantics keyword rows (selected): list of dicts
    #   {keyword, source, category, kw_type(SKW/MKW), match, broad_list, product,
    #    placement_mod, asp, acos_target}
    semantics_rows: list = field(default_factory=list)
    # Campaign rows (generated upstream or here): list of dicts of column->value
    campaign_rows: list = field(default_factory=list)


_EXT_REF_RE = re.compile(r"\[\d+\]")


def _sanitize_xludf(wb):
    """Clear vestigial Google-export helper formulas in the template's top/legend rows
    and drop the external-workbook link they create.

    Two problems they cause in Excel:
      • `_xludf.`/`__xludf.DUMMYFUNCTION` prefixes are read as references to an external
        workbook, and `'[1]SQP Report ...'` is a literal link to a second workbook that
        does not exist. Either one triggers 'couldn't get updated values from a linked
        workbook' and then BLOCKS recalculation (clean XLOOKUPs then show 0).
    These formulas are only legend/template artifacts in the top rows — clearing them is
    safe; the real data rows get fresh formulas written elsewhere."""
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 60)):
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                v = cell.value
                text = v.text if isinstance(v, ArrayFormula) else (v if isinstance(v, str) else None)
                if not text:
                    continue
                if ("_xludf" in text or "DUMMYFUNCTION" in text
                        or _EXT_REF_RE.search(text)):
                    cell.value = None
    # Drop the external-link package + <externalReferences> from workbook.xml.
    try:
        wb._external_links = []
    except Exception:  # noqa: BLE001
        pass


_SAMPLE_ASIN_RE = re.compile(r"B0[A-Z0-9]{8}", re.I)


def _strip_sample_notes(wb):
    """The shipped template carries sample-project ("Be Right Golf") text and
    sample ASINs in preamble / note / echo cells the per-sheet builders never
    touch — they only clear the data grid (rows at/after data_row). Blank those
    so one project's identity can never leak into another project's workbook.

    Two passes:
      1) A generic scrub of every managed sheet's preamble (rows above data_row):
         any plain-text cell holding an ASIN token or the sample brand name is a
         leftover filter echo / note, so it is cleared. Formulas are left alone,
         and the data grid itself is untouched. Builders that write ASIN-input
         lists run AFTER this and repopulate with the real project's ASINs.
      2) POE's "Customer Need" example list, which is plain product text (no ASIN)
         and so needs an explicit clear."""
    for title, lay in LAYOUTS.items():
        ws = wb[title] if title in wb.sheetnames else None
        if ws is None:
            continue
        for r in range(1, lay.data_row):                 # preamble only
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(row=r, column=c)
                v = cell.value
                if not isinstance(v, str) or v.startswith("="):
                    continue
                if "be right golf" in v.lower() or _SAMPLE_ASIN_RE.search(v):
                    cell.value = None
    if S_POE in wb.sheetnames:
        ws = wb[S_POE]
        for r in range(4, LAYOUTS[S_POE].data_row - 3):   # B4..B19, keep B20 header
            c = ws.cell(row=r, column=2)
            if isinstance(c.value, str):
                c.value = None


def build(inp: BuildInput, out_path: str) -> str:
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    _sanitize_xludf(wb)
    _strip_sample_notes(wb)

    has_str = bool(inp.str_table)
    sqp_tabs = _prepare_sqp_tabs(wb, inp.sqp_reports)
    _prepare_ba_tabs(wb, inp.ba_tst)

    _build_root_kw(wb[S_ROOT], inp)
    _build_asin_list(wb[S_ASIN], inp)
    _build_poe(wb[S_POE], inp)
    _build_h10(wb[S_H10], inp)
    _build_passthrough(wb, S_BRAND_ANALYTICS, inp.brand_analytics_table)
    _build_brand(wb[S_BRAND], inp)
    if has_str:
        _build_passthrough(wb, S_STR, inp.str_table)
    else:
        _clear(wb[S_STR], LAYOUTS[S_STR])
    _build_mkl(wb[S_MKL], inp)
    _build_semantics(wb[S_SEM], inp, has_str, sqp_tabs)
    _build_pat(wb[S_PAT], inp, has_str)
    _build_campaign(wb[S_CAMP], inp)

    # Force Excel to recompute on open. openpyxl writes formulas WITHOUT cached
    # results; if the file keeps the template's calcId, Excel assumes a compatible
    # engine already calculated it and shows the (missing) cached values as 0.
    # calcId=0 + fullCalcOnLoad + forceFullCalc makes Excel/Sheets recalculate.
    try:
        wb.calculation.calcId = 0
        wb.calculation.calcMode = "auto"
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except Exception:  # noqa: BLE001
        pass

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    return out_path


# --------------------------------------------------------------------------- #
# Per-sheet builders
# --------------------------------------------------------------------------- #
def _build_root_kw(ws, inp: BuildInput):
    """Write the client's Campaign Root KW categories into col A (A2 down)."""
    lay = LAYOUTS[S_ROOT]
    _clear(ws, lay)
    for i, cat in enumerate(inp.root_categories):
        _set(ws, "A", lay.data_row + i, cat)


def _build_asin_list(ws, inp: BuildInput):
    lay = LAYOUTS[S_ASIN]
    _clear(ws, lay)
    r = lay.data_row
    for p in inp.products:
        _set(ws, "A", r, p.get("sku", ""))
        _set(ws, "B", r, p.get("asin", ""))
        _set(ws, "C", r, p.get("name", ""))
        for col, f in _asin_formulas(r).items():
            _set(ws, col, r, f)
        r += 1
    # CTR data block (cols L..AA) — per-ASIN metrics so E/F/G SUMIFS resolve.
    cr = lay.data_row
    # CTR (S) and ACOS (Y) are derived, not pulled, and shown as live
    # percentages — a raw pulled ratio rendered as e.g. 0.002 / 1.7797 instead
    # of 0.2% / 177.97%. Kept as formulas so they always agree with the
    # Clicks/Impressions (R/Q) and Spend/Sales (T/V) in the same row.
    block = [("L", "asin"), ("M", "sku"), ("N", "title"), ("O", "state"),
             ("P", "profile"), ("Q", "impression"), ("R", "click"),
             ("T", "spend"), ("U", "cpc"), ("V", "sales"), ("W", "orders"),
             ("Z", "price"), ("AA", "asp")]
    for m in inp.asin_ctr:
        for col, key in block:
            v = m.get(key)
            if v not in (None, ""):
                _set(ws, col, cr, v)
        if m.get("click") not in (None, "") or m.get("impression") not in (None, ""):
            _setf(ws, "S", cr, f"=IF(Q{cr}=0,0,R{cr}/Q{cr})", PCT1)
        if m.get("sales") not in (None, "") or m.get("spend") not in (None, ""):
            _setf(ws, "Y", cr, f"=IF(V{cr}=0,0,T{cr}/V{cr})", PCT1)
        cr += 1


def _build_poe(ws, inp: BuildInput):
    lay = LAYOUTS[S_POE]
    _clear(ws, lay)
    r = lay.data_row
    needs, seen = [], set()          # distinct Customer Needs, in first-seen order
    # POE data table cols B..O passthrough, P formula, Q passthrough.
    # poe_tables values are rows of [B,C,D,E,F,G,H,I,J,K,L,M,N,O] (+ optional Q at idx14)
    for product, rows in inp.poe_tables.items():
        for row in rows:
            _write_rows(ws, [row[:14]], r, start_col=column_index_from_string("B"))
            _set(ws, "P", r, f"=ROUND(E{r}/12,0)")
            if len(row) > 14 and row[14] not in (None, ""):
                _set(ws, "Q", r, row[14])
            cn = str(row[0]).strip() if row else ""    # col B = Customer Need
            if cn and cn.lower() not in seen:
                seen.add(cn.lower())
                needs.append(cn)
            r += 1
    # Refill the top "Customer Need" summary list (B4..) that the sample-data strip
    # blanks. Populated from the upload's distinct Customer Needs so the block isn't
    # empty. This region is display-only (no formula references it), so the data
    # table and its Search Terms are untouched.
    sr, end = 4, lay.data_row - 4
    for cn in needs:
        if sr > end:
            break
        _set(ws, "B", sr, cn)
        sr += 1


def _build_h10(ws, inp: BuildInput):
    lay = LAYOUTS[S_H10]
    _clear(ws, lay)
    # ASIN input list A4.. (above the keyword header) — clear the template's
    # sample ASINs first so an unused slot never keeps a sample-project ASIN.
    for r in range(H10_ASIN_START, lay.header_row):
        _set(ws, "A", r, None)
    for i, asin in enumerate(inp.h10_asins[: lay.header_row - H10_ASIN_START]):
        _set(ws, "A", H10_ASIN_START + i, asin)
    # keyword table A15(header).. data 16+
    _write_rows(ws, inp.h10_table, lay.data_row, start_col=1)


def _build_brand(ws, inp: BuildInput):
    lay = LAYOUTS[S_BRAND]
    _clear(ws, lay)
    # Clear the template's sample ASIN(s) in the input list above the data row.
    for r in range(BRAND_ASIN_START, lay.header_row):
        _set(ws, "A", r, None)
    for i, asin in enumerate(inp.brand_asins[:1]):
        _set(ws, "A", BRAND_ASIN_START + i, asin)
    _write_rows(ws, inp.brand_table, lay.data_row, start_col=1)


def _build_passthrough(wb, sheet, rows):
    ws = wb[sheet]
    lay = LAYOUTS[sheet]
    _clear(ws, lay)
    _write_rows(ws, rows, lay.data_row, start_col=1)


def _build_mkl(ws, inp: BuildInput):
    lay = LAYOUTS[S_MKL]
    _clear(ws, lay)
    cols = {
        "AL": inp.own_branded_kws,
        "AM": inp.own_branded_searches,
        "AP": inp.competitor_kws,
        "AQ": inp.competitor_searches,
        "BD": inp.own_brand_asins,
        "BE": inp.main_competitor_asins,
        "BF": inp.lower_rated_asins,
        "BG": inp.higher_priced_asins,
        "BH": inp.bestselling_asins,
        "BI": inp.harvested_asins,
    }
    for col, vals in cols.items():
        for i, v in enumerate(vals):
            _set(ws, col, lay.data_row + i, v)


def _build_semantics(ws, inp: BuildInput, has_str, sqp_tabs):
    lay = LAYOUTS[S_SEM]
    _clear(ws, lay)
    r = lay.data_row
    for row in inp.semantics_rows:
        _set(ws, "A", r, row.get("keyword", ""))
        _set(ws, "B", r, row.get("source", ""))
        _set(ws, "K", r, row.get("category", ""))
        # L (KW Vol. / SKW-MKW), M (Match Type) and N (Broad KW List) are
        # user-controlled: EMPTY by default, written only from the user's manual
        # edits in the app (disp_* fields). The col-P campaign lookup keys off
        # L & M, so it shows "n/a" until the user fills them — by design.
        if row.get("disp_kw_type"):
            _set(ws, "M", r, row["disp_kw_type"])
        if row.get("disp_match"):
            _set(ws, "N", r, row["disp_match"])
        if row.get("disp_broad"):
            _set(ws, "O", r, row["disp_broad"])
        # H/I/J (Organic Rank, Impression Share, CTR) — also user-fillable.
        if row.get("organic_rank") not in (None, ""):
            _set(ws, "H", r, row["organic_rank"])
        if row.get("impression_share") not in (None, ""):
            _set(ws, "I", r, row["impression_share"])
        if row.get("ctr") not in (None, ""):
            _set(ws, "J", r, row["ctr"])
        if row.get("product"):
            _set(ws, "P", r, row["product"])
        if row.get("placement_mod") is not None:
            _setf(ws, "R", r, row["placement_mod"], PCT)        # placement modifier
        if row.get("asp") is not None:
            _setf(ws, "S", r, row["asp"], MONEY)                # ASP
        if row.get("acos_target") is not None:
            _setf(ws, "T", r, row["acos_target"], PCT)          # ACoS target
        for col, f in _sem_formulas(r, has_str, sqp_tabs).items():
            _set(ws, col, r, f)
        # number formats for the formula columns so they read naturally
        _set_fmt(ws, r, {"C": INT, "D": INT, "E": MONEY, "F": PCT1, "G": PCT1, "U": MONEY})
        r += 1


# PAT category -> placement bid adjustment (Product Pages). From the master sheet's
# S:T reference table. Cleared by the template build, so we re-write it here.
PAT_CATEGORY_ADJ = [("Low Rated", 0.5), ("High Priced", 0.35),
                    ("Bestselling", 0.1), ("Main", 0.25)]


def _build_pat(ws, inp: BuildInput, has_str):
    lay = LAYOUTS[S_PAT]
    _clear(ws, lay)
    # S:T reference table (category -> placement bid adjustment), whole-column
    # XLOOKUP target for the M formula. Lives in cols S/T regardless of row count.
    for i, (name, adj) in enumerate(PAT_CATEGORY_ADJ):
        _set(ws, "S", lay.data_row + i, name)
        _setf(ws, "T", lay.data_row + i, adj, PCT)
    # One row per competitor ASIN we target.
    r = lay.data_row
    for t in inp.pat_targets:
        _set(ws, "A", r, t.get("source", "Competitor"))
        _set(ws, "B", r, t.get("asin", ""))
        if t.get("title"):
            _set(ws, "C", r, t["title"])
        if t.get("type"):
            _set(ws, "J", r, t["type"])
        if t.get("product"):
            _set(ws, "K", r, t["product"])
        if t.get("asp") is not None:
            _setf(ws, "N", r, t["asp"], MONEY)
        if t.get("acos") is not None:
            _setf(ws, "O", r, t["acos"], PCT)
        for col, f in _pat_formulas(r, has_str).items():
            _set(ws, col, r, f)
        # Conversion Rate (H): PAT targets are competitor ASINs (no per-ASIN CVR in
        # any upload), so use the POE-derived product CVR passed in. Overrides the
        # STR-by-ASIN formula, which never matched and left it blank.
        if t.get("cvr") not in (None, ""):
            _setf(ws, "H", r, t["cvr"], PCT1)
        _set_fmt(ws, r, {"E": INT, "F": MONEY, "G": PCT1, "H": PCT1, "M": PCT, "Q": MONEY})
        r += 1


def _build_campaign(ws, inp: BuildInput):
    lay = LAYOUTS[S_CAMP]
    _clear(ws, lay)
    r = lay.data_row
    for row in inp.campaign_rows:
        for col, val in row.items():
            _set(ws, col, r, val)
        for col, f in _camp_formulas(r, inp.brand).items():
            # do not overwrite an explicitly provided value
            if col not in row:
                _set(ws, col, r, f)
        # percent / money formats for the numeric campaign columns
        _set_fmt(ws, r, {"Q": PCT, "R": PCT, "S": PCT, "Z": MONEY,
                         "AH": PCT, "AI": MONEY, "AJ": PCT, "AK": PCT, "AL": MONEY})
        r += 1


# --------------------------------------------------------------------------- #
# SQP / BA TST tab management
# --------------------------------------------------------------------------- #
# Excel forbids these in sheet names and caps length at 31. A name over 31 chars
# makes Excel "repair" the tab but leaves every formula referencing the long name
# as a BROKEN external link -> linked-workbook banner -> recalc blocked.
_BAD_TITLE = re.compile(r"[\\/?*\[\]:]")


def _safe_sheet_title(name, existing):
    t = _BAD_TITLE.sub("-", str(name or "Sheet")).strip() or "Sheet"
    t = t[:31]
    if t not in existing:
        return t
    # ensure uniqueness within the 31-char budget
    i = 2
    while True:
        suffix = f" {i}"
        cand = t[:31 - len(suffix)] + suffix
        if cand not in existing:
            return cand
        i += 1


def _prepare_sqp_tabs(wb, sqp_reports: dict):
    """Rename/clone the template SQP tab to one per user ASIN. Returns tab names
    (each a valid, <=31-char title that the Semantics formula will reference)."""
    base = wb[TEMPLATE_SQP]
    lay = LAYOUTS[TEMPLATE_SQP]
    asins = list(sqp_reports.keys())
    tabs = []

    def _retarget(ws, asin):
        """Point the tab's sample ASIN echo (A10) at this project's ASIN so the
        template's sample ASIN never survives."""
        _set(ws, "D", 4, asin)
        _set(ws, "A", 10, f'ASIN or Product=["{asin}"]' if asin else None)

    if not asins:
        # No SQP provided: clear the example data AND strip the sample ASIN from
        # the tab name / echo so the template's ASIN doesn't leak.
        _clear(base, lay)
        _retarget(base, "")
        existing = {ws.title for ws in wb.worksheets if ws is not base}
        base.title = _safe_sheet_title("SQP Report", existing)
        return []
    existing = {ws.title for ws in wb.worksheets if ws is not base}
    # first ASIN reuses the base tab
    first = asins[0]
    base.title = _safe_sheet_title(f"SQP Report {first}", existing)
    existing.add(base.title)
    _clear(base, lay)
    _retarget(base, first)
    _write_rows(base, sqp_reports[first], lay.data_row, start_col=1)
    tabs.append(base.title)
    # additional ASINs clone the base
    for asin in asins[1:3]:
        ws = wb.copy_worksheet(base)
        ws.title = _safe_sheet_title(f"SQP Report {asin}", existing)
        existing.add(ws.title)
        _clear(ws, lay)
        _retarget(ws, asin)
        _write_rows(ws, sqp_reports[asin], lay.data_row, start_col=1)
        tabs.append(ws.title)
    return tabs


def _prepare_ba_tabs(wb, ba_tst: dict):
    """Repurpose the template BA TST tabs to THIS project's uploaded words.

    The template ships example BA TST tabs named after a sample project (e.g.
    "BA TST - Golf"). The old logic only filled a tab whose name already matched
    an uploaded word, so a project with different words kept the sample tab names
    and lost its own data — surfacing another project's identity in the workbook.
    Instead, rename one tab per uploaded word (cloning when there are more words
    than template tabs) and delete any leftover example tabs — the same approach
    as the SQP step."""
    base_tabs = [ws for ws in wb.worksheets if ws.title.startswith("BA TST")]
    if not base_tabs:
        return
    words = [w for w in ba_tst.keys() if str(w).strip()]
    if not words:
        # No BA TST uploaded: drop the sample tabs entirely (nothing references
        # them by formula) so none of their sample identity survives.
        for ws in base_tabs:
            wb.remove(ws)
        return
    existing = {ws.title for ws in wb.worksheets}
    base = base_tabs[0]
    for i, word in enumerate(words):
        ws = base_tabs[i] if i < len(base_tabs) else wb.copy_worksheet(base)
        _reset_ba_tab(ws, word)
        existing.discard(ws.title)
        ws.title = _safe_sheet_title(f"BA TST - {word}", existing)
        existing.add(ws.title)
        data = ba_tst[word]
        # New format carries the upload's own header (Selection is col A);
        # write it over the template's fixed header so columns line up. Legacy
        # (list) values fall back to the template header.
        if isinstance(data, dict):
            header, rows = data.get("header") or [], data.get("rows") or []
        else:
            header, rows = None, data
        if header:
            for c in range(1, ws.max_column + 1):       # clear the template header
                _put(ws, BA_LAYOUT.header_row, c, None)
            _write_rows(ws, [header], BA_LAYOUT.header_row, start_col=1)
        _write_rows(ws, rows, BA_LAYOUT.data_row, start_col=1)
    # Remove leftover template example tabs (e.g. the sample "Golf" set).
    for ws in base_tabs[len(words):]:
        wb.remove(ws)


def _reset_ba_tab(ws, word):
    """Wipe a template BA TST tab of all sample content before it is reused for a
    new word. `_clear` alone is not enough: sample data ran past the managed
    columns, and the control panel (rows 1..8) carries the old filter word — both
    would otherwise leak the sample project's identity."""
    # Data rows, FULL width (sample data extended past the managed last column).
    end = min(BA_LAYOUT.clear_to, ws.max_row)
    for r in range(BA_LAYOUT.data_row, end + 1):
        for c in range(1, ws.max_column + 1):
            _put(ws, r, c, None)
    # Control panel: retarget the filter-word cells to this project's word.
    # Label-driven (not word-driven) so it is independent of the template samples.
    for r in range(1, BA_LAYOUT.header_row):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if not isinstance(v, str):
                continue
            if v.strip().lower() == "search term filter word":
                ws.cell(row=r, column=c + 1).value = word
            elif re.match(r'\s*search\s*term\s*=\s*\[', v, re.I):
                ws.cell(row=r, column=c).value = f'Search Term=["{word}"]'
