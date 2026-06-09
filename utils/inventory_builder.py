"""Build the 4-sheet FBA Inventory Transfer workbook exactly per the skill spec.

Sheet order in the workbook (and BUILD order is 2,3,4 then 1 last so the Sheet-1
cross-sheet SUMIFS resolve):
  1. Inventory Transfer <Market>   (formula-driven working sheet — built last)
  2. Business Reports              (23 cols A-W, Legend in W)
  3. Manage FBA Inventory          (22 cols A-V, sku in A)
  4. All Listings Report           (20 reference cols)

Every numeric cell in Sheet 1 cols M-Y is a formula referencing Sheets 2 & 3 —
no hardcoded numbers, exactly as the skill requires.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ----------------------------------------------------------------- styling ---
NAVY = "1F2D3D"
WHITE = "FFFFFF"
FONT = "Arial"

BAND_ROW_FILL = {
    "BAND A": ("D6E4F0", "EBF4FB"),
    "BAND B": ("D5F5E3", "EAFAF1"),
    "BAND C": ("FDEBD0", "FEF5E7"),
    "EOL":    ("FADBD8", "FDFEFE"),
}
BAND_PILL = {
    "BAND A": "3B82F6",
    "BAND B": "22C55E",
    "BAND C": "D4A017",
    "EOL":    "E74C3C",
}

_thin = Side(style="thin", color="D0D0D0")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

PERIOD_LABELS = ["L30D", "L60D", "L90D", "L180D", "L270D", "L365D"]

# Sheet 2 — Business Reports headers (23, A-W). Order is load-bearing.
BR_HEADERS = [
    "(Parent) ASIN", "(Child) ASIN", "Title", "SKU",
    "Sessions - Total", "Sessions - Total - B2B",
    "Session Percentage - Total", "Session Percentage - Total - B2B",
    "Page Views - Total", "Page Views - Total - B2B",
    "Page Views Percentage - Total", "Page Views Percentage - Total - B2B",
    "Featured Offer (Buy Box) Percentage", "Featured Offer (Buy Box) Percentage - B2B",
    "Units Ordered", "Units Ordered - B2B",
    "Unit Session Percentage", "Unit Session Percentage - B2B",
    "Ordered Product Sales", "Ordered Product Sales - B2B",
    "Total Order Items", "Total Order Items - B2B",
    "Legend",
]

# Sheet 3 — Manage FBA Inventory headers (22, A-V) = raw report column names.
FBA_HEADERS = [
    "sku", "fnsku", "asin", "product-name", "condition", "your-price",
    "mfn-listing-exists", "mfn-fulfillable-quantity", "afn-listing-exists",
    "afn-warehouse-quantity", "afn-fulfillable-quantity", "afn-unsellable-quantity",
    "afn-reserved-quantity", "afn-total-quantity", "per-unit-volume",
    "afn-inbound-working-quantity", "afn-inbound-shipped-quantity",
    "afn-inbound-receiving-quantity", "afn-researching-quantity",
    "afn-reserved-future-supply", "afn-future-supply-buyable", "store",
]

# Sheet 4 — All Listings Report headers (20).
LISTING_HEADERS = [
    "asin1", "seller-sku", "item-name", "price", "quantity", "open-date",
    "product-id-type", "item-note", "item-condition", "will-ship-internationally",
    "expedited-shipping", "product-id", "pending-quantity", "fulfillment-channel",
    "merchant-shipping-group", "standard-price-point", "ProductTaxCode", "status",
    "minimum-seller-allowed-price", "maximum-seller-allowed-price",
]

# Sheet 1 — 22 column headers (A-V), matching the reference working file exactly
# (no Parent SKU/ASIN/Title columns).
S1_HEADERS = [
    "Child SKU", "Child ASIN", "Child Title", "Brand", "Product Category",
    "Size", "Color", "Fulfillment", "BAND", "L30D", "L60D", "L90D", "L180D",
    "L270D", "L365D", "Avg.", "TW Inventory", "Working", "Intransit",
    "Reserved/FC Transfer", "Total Inventory", "Desired Replenishment",
]

_NUMERIC_FBA_COLS = set(range(8, 22))  # mfn-fulfillable.. onward are numbers


def _to_num(val):
    if val is None or val == "":
        return None
    try:
        f = float(str(val).replace(",", "").replace("$", "").strip())
        return int(f) if f.is_integer() else f
    except (ValueError, AttributeError):
        return val


def _hdr_cell(cell, size=10, fill=NAVY, color=WHITE, bold=True):
    cell.font = Font(name=FONT, size=size, bold=bold, color=color)
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER


def _plain(cell, size=10, bold=False):
    cell.font = Font(name=FONT, size=size, bold=bold)
    cell.border = BORDER


def build_inventory_workbook(reports, product_list, output_path, market="US",
                             increase_factor=0.0):
    """Build and save the workbook. Returns output_path.

    reports: spapi_client.pull_inventory_reports output.
    product_list: band_classifier.build_product_list output (possibly user-edited).
    increase_factor: replenishment safety buffer (0 = none, 0.1 = +10%) -> cell P5.
    """
    s1_name = f"Inventory Transfer {market}"

    wb = Workbook()
    wb.remove(wb.active)

    # Build sheets 2,3,4 FIRST.
    _build_business_reports(wb, reports, product_list, market)
    _build_manage_fba(wb, reports, market)
    _build_all_listings(wb, reports, market)
    # Sheet 1 LAST, then move it to the front.
    _build_sheet1(wb, product_list, s1_name, market, increase_factor)
    wb.move_sheet(s1_name, -(len(wb.sheetnames) - 1))

    wb.save(output_path)
    return output_path


# ---------------------------------------------------- Sheet 2: Business Reports ---

def _build_business_reports(wb, reports, product_list, market):
    ws = wb.create_sheet("Business Reports")
    ws["A1"] = market
    ws["A2"] = "Reports > Business Reports > Detail Page Sales and Traffic"
    for i, h in enumerate(BR_HEADERS, start=1):
        _hdr_cell(ws.cell(row=3, column=i, value=h))

    sales_meta = reports.get("sales_meta", {})
    # asin -> (sku, title) for filling cols C/D
    asin_info = {p["asin"]: (p["sku"], p["title"]) for p in product_list if p.get("asin")}
    # Unique child ASINs (dedupe so per-ASIN sales aren't double-counted by Sheet 1).
    asins = list(dict.fromkeys(p["asin"] for p in product_list if p.get("asin")))

    # One canonical parent ASIN per child across all periods (Amazon sometimes
    # reports different parents per period — pick the most frequent, prefer a real
    # non-self parent on ties).
    parent_votes = {}
    for label in PERIOD_LABELS:
        for child, entry in sales_meta.get(label, {}).items():
            p = (entry or {}).get("parentAsin")
            if p:
                parent_votes.setdefault(child, {}).setdefault(p, 0)
                parent_votes[child][p] += 1

    def canonical_parent(child):
        votes = parent_votes.get(child)
        if not votes:
            return child  # single-ASIN product with no sales data -> itself
        return max(votes, key=lambda k: (votes[k], k != child))

    row = 4
    for asin in asins:
        sku, title = asin_info.get(asin, ("", ""))
        parent = canonical_parent(asin)
        for label in PERIOD_LABELS:
            entry = sales_meta.get(label, {}).get(asin, {})
            sales = entry.get("salesByAsin", {}) or {}
            traffic = entry.get("trafficByAsin", {}) or {}
            ops = (sales.get("orderedProductSales") or {}).get("amount", 0)
            ops_b2b = (sales.get("orderedProductSalesB2B") or {}).get("amount", 0)
            values = [
                parent,                                           # A canonical parent
                asin,                                             # B  <- SUMIFS key
                title,                                            # C
                sku,                                              # D
                traffic.get("sessions", 0),                       # E
                traffic.get("sessionsB2B", 0),                    # F
                traffic.get("sessionPercentage", 0),              # G
                traffic.get("sessionPercentageB2B", 0),           # H
                traffic.get("pageViews", 0),                      # I
                traffic.get("pageViewsB2B", 0),                   # J
                traffic.get("pageViewsPercentage", 0),            # K
                traffic.get("pageViewsPercentageB2B", 0),         # L
                traffic.get("buyBoxPercentage", 0),               # M
                traffic.get("buyBoxPercentageB2B", 0),            # N
                int(sales.get("unitsOrdered", 0) or 0),           # O  <- SUMIFS return
                int(sales.get("unitsOrderedB2B", 0) or 0),        # P
                traffic.get("unitSessionPercentage", 0),          # Q
                traffic.get("unitSessionPercentageB2B", 0),       # R
                ops,                                              # S
                ops_b2b,                                          # T
                int(sales.get("totalOrderItems", 0) or 0),        # U
                int(sales.get("totalOrderItemsB2B", 0) or 0),     # V
                label,                                            # W  <- Legend
            ]
            for i, v in enumerate(values, start=1):
                c = ws.cell(row=row, column=i, value=v)
                _plain(c)
            row += 1
    _autosize(ws, BR_HEADERS)


# -------------------------------------------------- Sheet 3: Manage FBA Inventory ---

def _build_manage_fba(wb, reports, market):
    ws = wb.create_sheet("Manage FBA Inventory")
    ws["A1"] = market
    ws["A2"] = "Reports > Fulfillment > Manage FBA Inventory"
    for i, h in enumerate(FBA_HEADERS, start=1):
        _hdr_cell(ws.cell(row=3, column=i, value=h))

    row = 4
    for r in reports.get("fba_inventory", []):
        for i, key in enumerate(FBA_HEADERS, start=1):
            val = r.get(key, "")
            if i in _NUMERIC_FBA_COLS:
                val = _to_num(val) or 0
            c = ws.cell(row=row, column=i, value=val)
            _plain(c)
        row += 1
    _autosize(ws, FBA_HEADERS)


# ----------------------------------------------------- Sheet 4: All Listings ---

def _build_all_listings(wb, reports, market):
    ws = wb.create_sheet("All Listings Report")
    ws["A1"] = market
    ws["A2"] = "Reports > Inventory Reports > All Listings Report"
    for i, h in enumerate(LISTING_HEADERS, start=1):
        _hdr_cell(ws.cell(row=3, column=i, value=h))

    row = 4
    for r in reports.get("open_listings", []):
        for i, key in enumerate(LISTING_HEADERS, start=1):
            c = ws.cell(row=row, column=i, value=r.get(key, ""))
            _plain(c)
        row += 1
    _autosize(ws, LISTING_HEADERS)


# ------------------------------------------------ Sheet 1: Inventory Transfer ---

def _build_sheet1(wb, product_list, s1_name, market, increase_factor=0.0):
    ws = wb.create_sheet(s1_name)
    n = len(product_list)
    first_data = 11
    last_data = first_data + n - 1 if n else first_data
    sub_bottom = max(100, last_data)  # subtotal/formula range bottom

    # Period divisors row 5 (J..O) and increase factor P5 — matches reference.
    for col, val in zip(["J", "K", "L", "M", "N", "O"], [1, 2, 3, 6, 9, 12]):
        ws[f"{col}5"] = val
        ws[f"{col}5"].font = Font(name=FONT, size=9, italic=True, color="888888")
    ws["P5"] = round(float(increase_factor or 0), 4)
    ws["J6"] = "Sales"
    ws["J6"].font = Font(name=FONT, size=9, italic=True, color="888888")
    ws["P4"] = "Increase"
    ws["P4"].font = Font(name=FONT, size=9, italic=True, color="888888")

    _build_band_summary(ws)
    _build_subtotals(ws, sub_bottom)

    # Column headers row 10.
    for i, h in enumerate(S1_HEADERS, start=1):
        _hdr_cell(ws.cell(row=10, column=i, value=h))

    # Data rows.
    for idx, p in enumerate(product_list):
        row = first_data + idx
        _write_product_row(ws, row, p, s1_name)

    # Freeze rows 1-10.
    ws.freeze_panes = "A11"
    _format_sheet1_columns(ws)


def _build_band_summary(ws):
    """Band reference table in cols T-Z (20-26), rows 1-6 — matches reference.

    Column references inside: I=BAND, U=Total Inventory, V=Desired Replenishment.
    """
    headers = ["BAND", "Desired MoS", "Total Count of ASINs", "Instock ASINs",
               "Instock Rate", "Count of ASINs under Restock", "Desired Replenishment"]
    for j, h in enumerate(headers):
        _hdr_cell(ws.cell(row=1, column=20 + j, value=h), size=9)

    bands = [("BAND A", 3), ("BAND B", 2.5), ("BAND C", 2), ("EOL", 0)]
    for k, (label, mos) in enumerate(bands):
        r = 2 + k
        ws.cell(row=r, column=20, value=label).font = Font(name=FONT, size=9, bold=True)  # T
        ws.cell(row=r, column=21, value=mos).font = Font(name=FONT, size=9)               # U
        ws.cell(row=r, column=22, value=f"=COUNTIFS(I:I,T{r})")                            # V Total
        ws.cell(row=r, column=23, value=f'=COUNTIFS(I:I,T{r},U:U,">0")')                   # W Instock
        ws.cell(row=r, column=24, value=f"=IFERROR(W{r}/V{r},0)")                          # X Rate
        ws.cell(row=r, column=25, value=f'=COUNTIFS(I:I,T{r},V:V,">0")')                   # Y Restock
        ws.cell(row=r, column=26, value=f"=SUMIFS(V:V,$I:$I,$T{r})")                       # Z Desired
        ws.cell(row=r, column=24).number_format = "0%"

    # Totals row 6.
    ws.cell(row=6, column=20, value="TOTALS").font = Font(name=FONT, size=9, bold=True)
    ws.cell(row=6, column=22, value="=SUM(V2:V5)")
    ws.cell(row=6, column=23, value="=SUM(W2:W5)")
    ws.cell(row=6, column=24, value="=IFERROR(W6/V6,0)")
    ws.cell(row=6, column=24).number_format = "0%"
    ws.cell(row=6, column=25, value="=SUM(Y2:Y5)")
    ws.cell(row=6, column=26, value="=SUM(Z2:Z5)")


def _build_subtotals(ws, bottom):
    ws["I9"] = "Subtotals"
    ws["I9"].font = Font(name=FONT, size=10, bold=True)
    # Sales J-O and inventory Q-V (P=Avg skipped), matching reference.
    for col in ["J", "K", "L", "M", "N", "O", "Q", "R", "S", "T", "U", "V"]:
        ws[f"{col}9"] = f"=SUBTOTAL(9,{col}{11}:{col}{bottom})"
        ws[f"{col}9"].font = Font(name=FONT, size=10, bold=True)


def _write_product_row(ws, row, p, s1_name):
    band = p.get("band") or ""

    # A-H static fields (no parent columns in the reference layout). Minimal styling:
    # plain cells, no row fills.
    static = [
        p.get("sku", ""), p.get("asin", ""), p.get("title", ""),
        p.get("brand", ""), p.get("category", ""), p.get("size", ""),
        p.get("color", ""), p.get("fulfillment", ""),
    ]
    for i, v in enumerate(static, start=1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = Font(name=FONT, size=10)
        c.border = BORDER

    # BAND (col I = 9) — plain bold text, no coloured pill.
    lc = ws.cell(row=row, column=9, value=band)
    lc.font = Font(name=FONT, size=10, bold=True)
    lc.alignment = Alignment(horizontal="center", vertical="center")
    lc.border = BORDER

    b = f"$B{row}"           # child ASIN ref (sales SUMIFS key, col B)
    a = f"A{row}"            # child SKU ref (inventory SUMIFS key, col A)
    BR = "'Business Reports'"
    FBA = "'Manage FBA Inventory'"
    s1 = f"'{s1_name}'"

    formulas = {
        10: f"=SUMIFS({BR}!$O:$O,{BR}!$B:$B,{b},{BR}!$W:$W,\"L30D\")",   # J L30D
        11: f"=SUMIFS({BR}!$O:$O,{BR}!$B:$B,{b},{BR}!$W:$W,\"L60D\")",   # K L60D
        12: f"=SUMIFS({BR}!$O:$O,{BR}!$B:$B,{b},{BR}!$W:$W,\"L90D\")",   # L L90D
        13: f"=SUMIFS({BR}!$O:$O,{BR}!$B:$B,{b},{BR}!$W:$W,\"L180D\")",  # M L180D
        14: f"=SUMIFS({BR}!$O:$O,{BR}!$B:$B,{b},{BR}!$W:$W,\"L270D\")",  # N L270D
        15: f"=SUMIFS({BR}!$O:$O,{BR}!$B:$B,{b},{BR}!$W:$W,\"L365D\")",  # O L365D
        16: (f"=ROUND(AVERAGE(J{row}/J$5,K{row}/K$5,L{row}/L$5,"
             f"M{row}/M$5,N{row}/N$5,O{row}/O$5),1)"),                   # P Avg
        17: f"=SUMIFS({FBA}!K:K,{FBA}!A:A,{s1}!{a})",                    # Q TW Inventory
        18: f"=SUMIFS({FBA}!P:P,{FBA}!A:A,{s1}!{a})",                    # R Working
        19: (f"=SUM(SUMIFS({FBA}!Q:Q,{FBA}!A:A,{s1}!{a}),"
             f"SUMIFS({FBA}!R:R,{FBA}!A:A,{s1}!{a}))"),                  # S Intransit
        20: (f"=SUM(SUMIFS({FBA}!M:M,{FBA}!A:A,{s1}!{a}),"
             f"SUMIFS({FBA}!S:S,{FBA}!A:A,{s1}!{a}))"),                  # T Reserved/FC
        21: f"=SUM(Q{row},R{row},S{row},T{row})",                       # U Total Inv
        22: (f"=IFERROR(MAX(0,IF($I{row}=$T$2,ROUNDUP(P{row}*(1+$P$5)*$U$2,0),"
             f"IF($I{row}=$T$3,ROUNDUP(P{row}*(1+$P$5)*$U$3,0),"
             f"IF($I{row}=$T$4,ROUNDUP(P{row}*(1+$P$5)*$U$4,0),"
             f"ROUNDUP(P{row}*(1+$P$5)*$U$5,0))))-U{row}),0)"),         # V Desired Repl
    }
    for col, f in formulas.items():
        c = ws.cell(row=row, column=col, value=f)
        c.font = Font(name=FONT, size=10)
        c.border = BORDER


def _format_sheet1_columns(ws):
    widths = {"A": 16, "B": 14, "C": 30, "D": 14, "E": 16, "F": 8, "G": 10,
              "H": 12, "I": 9}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    for col in ["J", "K", "L", "M", "N", "O", "Q", "R", "S", "T", "U", "V"]:
        for row in range(9, 201):
            ws[f"{col}{row}"].number_format = "#,##0"
    for row in range(9, 201):
        ws[f"P{row}"].number_format = "0.0"


def _autosize(ws, headers, cap=40):
    for i, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = min(cap, max(10, len(str(h)) + 2))
